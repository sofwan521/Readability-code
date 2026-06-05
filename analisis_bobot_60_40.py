"""
Moodle Readability Analyzer v3.0
==================================
Menganalisis readability kode PHP di seluruh folder Moodle.
Menggunakan lizard untuk ekstraksi fungsi yang akurat.

Usage:
  # Analisis SEMUA folder di moodle/public (TANPA sampling — semua file diproses)
  python moodle_readability_analyzer.py --path "C:/path/to/moodle/public"

  # Analisis modul tertentu saja (dengan sampling 100 file per modul)
  python moodle_readability_analyzer.py --path "C:/path/to/moodle/public" --modules lib mod search --sample 100

  # Analisis modul tertentu, tanpa sampling
  python moodle_readability_analyzer.py --path "C:/path/to/moodle/public" --modules lib mod
"""

import sys
import random
import argparse
from pathlib import Path
from dataclasses import dataclass
from typing import List, Optional, Tuple

# ── Auto-install dependencies ────────────────────────────────────────────────
def install(pkg):
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

try:
    import pandas as pd
except ImportError:
    print("Installing pandas..."); install("pandas"); import pandas as pd

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
except ImportError:
    print("Installing matplotlib..."); install("matplotlib")
    import matplotlib; matplotlib.use('Agg')
    import matplotlib.pyplot as plt

try:
    import seaborn as sns
except ImportError:
    print("Installing seaborn..."); install("seaborn"); import seaborn as sns

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("Installing openpyxl..."); install("openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

try:
    import lizard
except ImportError:
    print("Installing lizard..."); install("lizard"); import lizard

plt.style.use('seaborn-v0_8-darkgrid')
sns.set_palette("husl")
# ─────────────────────────────────────────────────────────────────────────────


# ── Folder & file yang di-skip ───────────────────────────────────────────────
IGNORE_FOLDERS = {
    'vendor', 'node_modules', 'tests', 'yui', 'amd',
    'cache', 'local', 'behat', 'fixtures', 'coverage',
    'pix', 'lang', 'theme', 'userpix', '.git'
}
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class FunctionMetrics:
    module: str
    file_path: str
    file_name: str
    function_name: str
    start_line: int
    end_line: int
    total_lines: int
    code_lines: int
    comment_lines: int
    blank_lines: int
    comment_density: float        # dalam %
    function_length_score: float  # 0-100
    comment_density_score: float  # 0-100
    readability_score: float      # 0-100
    category: str                 # High / Medium / Low


class MoodleReadabilityAnalyzer:
    # ── Konstanta ────────────────────────────────────────────────────────────
    IDEAL_LOC = 20
    MAX_LOC   = 100
    IDEAL_CD  = 0.25   # 25%
    W_LOC     = 0.6
    W_CD      = 0.4
    # ─────────────────────────────────────────────────────────────────────────

    def __init__(
        self,
        moodle_path: str,
        modules: Optional[List[str]] = None,
        sample_size: Optional[int] = None,
        seed: int = 42,
    ):
        self.moodle_path = Path(moodle_path)
        self.modules     = modules
        self.sample_size = sample_size
        self.seed        = seed
        self.results: List[FunctionMetrics] = []

        if not self.moodle_path.exists():
            raise FileNotFoundError(f"Path tidak ditemukan: {moodle_path}")

    # ── Helper: apakah file harus di-skip? ──────────────────────────────────
    def _should_skip(self, php_file: Path) -> bool:
        parts = str(php_file).lower().replace('\\', '/').split('/')
        for part in parts:
            if part in IGNORE_FOLDERS:
                return True
            if part.endswith('.php') and ('_test' in part or part.startswith('test_')):
                return True
        return False

    # ── Helper: hitung komentar & baris kosong ──────────────────────────────
    def _comment_metrics(self, file_path: Path, start: int, end: int) -> Tuple[int, int]:
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
        except Exception:
            return 0, 0

        func_lines = lines[start - 1: end]
        comment_lines = blank_lines = 0
        in_block = False

        for line in func_lines:
            s = line.strip()
            if not s:
                blank_lines += 1
                continue
            if '/*' in s:
                in_block = True
                comment_lines += 1
                if '*/' in s:
                    in_block = False
                continue
            if in_block:
                comment_lines += 1
                if '*/' in s:
                    in_block = False
                continue
            if s.startswith(('//', '#', '*')):
                comment_lines += 1
                continue
            if '//' in s or '#' in s:
                comment_lines += 1

        return comment_lines, blank_lines

    # ── Skor ─────────────────────────────────────────────────────────────────
    def _score_loc(self, loc: int) -> float:
        if loc <= self.IDEAL_LOC: return 100.0
        if loc >= self.MAX_LOC:   return 0.0
        return max(0.0, 100 - (loc - self.IDEAL_LOC) / (self.MAX_LOC - self.IDEAL_LOC) * 100)

    def _score_cd(self, cd: float) -> float:
        if cd >= self.IDEAL_CD: return 100.0
        return min(100.0, cd / self.IDEAL_CD * 100)

    def _readability(self, s_loc: float, s_cd: float) -> float:
        return self.W_LOC * s_loc + self.W_CD * s_cd

    def _category(self, score: float) -> str:
        if score >= 80: return "High"
        if score >= 50: return "Medium"
        return "Low"

    # ── Analisis satu file ───────────────────────────────────────────────────
    def _analyze_file(self, file_path: Path, module: str) -> int:
        try:
            analysis = lizard.analyze_file(str(file_path))
        except Exception:
            return 0

        count = 0
        for func in analysis.function_list:
            loc   = func.nloc
            start = func.start_line
            end   = getattr(func, 'end_line', start + loc)

            if loc == 0:
                continue

            comments, blanks = self._comment_metrics(file_path, start, end)
            total_loc = loc + comments
            cd        = comments / total_loc if total_loc > 0 else 0.0

            s_loc   = self._score_loc(loc)
            s_cd    = self._score_cd(cd)
            r_score = self._readability(s_loc, s_cd)

            self.results.append(FunctionMetrics(
                module                = module,
                file_path             = str(file_path.relative_to(self.moodle_path)),
                file_name             = file_path.name,
                function_name         = func.name,
                start_line            = start,
                end_line              = end,
                total_lines           = end - start + 1,
                code_lines            = loc,
                comment_lines         = comments,
                blank_lines           = blanks,
                comment_density       = round(cd * 100, 2),
                function_length_score = round(s_loc, 2),
                comment_density_score = round(s_cd, 2),
                readability_score     = round(r_score, 2),
                category              = self._category(r_score),
            ))
            count += 1
        return count

    # ── Tentukan daftar modul ────────────────────────────────────────────────
    def _resolve_modules(self) -> List[str]:
        """
        Jika modules tidak ditentukan, ambil SEMUA subfolder langsung
        di moodle_path yang bukan IGNORE_FOLDERS.
        """
        if self.modules:
            return self.modules

        folders = []
        for item in sorted(self.moodle_path.iterdir()):
            if item.is_dir() and item.name.lower() not in IGNORE_FOLDERS:
                folders.append(item.name)
        return folders

    # ── Main: analisis semua modul ───────────────────────────────────────────
    def analyze(self) -> None:
        modules = self._resolve_modules()

        # Jika mode "semua modul" (--modules tidak diberikan), nonaktifkan sampling
        all_modules_mode = not self.modules
        effective_sample = None if all_modules_mode else self.sample_size

        print("=" * 60)
        print("  MOODLE READABILITY ANALYZER v3.0")
        print("=" * 60)
        print(f"Path   : {self.moodle_path}")
        print(f"Modul  : {', '.join(modules)}")
        if all_modules_mode:
            print(f"Sample : semua file (sampling dinonaktifkan — mode semua modul)")
        else:
            print(f"Sample : {effective_sample if effective_sample else 'semua file'} per modul")
        print(f"Seed   : {self.seed}")
        print("=" * 60 + "\n")

        grand_files = grand_funcs = 0

        for module in modules:
            module_path = self.moodle_path / module
            if not module_path.exists():
                print(f"[SKIP] Modul '{module}' tidak ditemukan.")
                continue

            all_files   = list(module_path.rglob("*.php"))
            valid_files = [f for f in all_files if not self._should_skip(f)]
            skipped     = len(all_files) - len(valid_files)

            if effective_sample and len(valid_files) > effective_sample:
                random.seed(self.seed)
                sampled     = random.sample(valid_files, effective_sample)
                sample_info = f"{effective_sample} (acak dari {len(valid_files)})"
            else:
                sampled     = valid_files
                sample_info = f"{len(sampled)} (semua)"

            print(f"[{module}]")
            print(f"  Total PHP     : {len(all_files)}")
            print(f"  Setelah filter: {len(valid_files)}  ({skipped} di-skip)")
            print(f"  Dianalisis    : {sample_info}")

            funcs_before = len(self.results)
            for php_file in sampled:
                self._analyze_file(php_file, module)

            found = len(self.results) - funcs_before
            grand_files += len(sampled)
            grand_funcs += found
            print(f"  Fungsi ditemukan: {found}\n")

        print("=" * 60)
        print(f"  TOTAL file dianalisis : {grand_files}")
        print(f"  TOTAL fungsi ditemukan: {grand_funcs}")
        print("=" * 60 + "\n")

    # ── Export Excel ─────────────────────────────────────────────────────────
    def export_excel(self, output_file: str) -> None:
        print(f"Mengekspor ke: {output_file}")

        df = pd.DataFrame([{
            'Module'                : r.module,
            'File'                  : r.file_name,
            'File Path'             : r.file_path,
            'Function'              : r.function_name,
            'Start Line'            : r.start_line,
            'End Line'              : r.end_line,
            'Total Lines'           : r.total_lines,
            'Code Lines (LOC)'      : r.code_lines,
            'Comment Lines'         : r.comment_lines,
            'Blank Lines'           : r.blank_lines,
            'Comment Density (%)'   : r.comment_density,
            'Function Length Score' : r.function_length_score,
            'Comment Density Score' : r.comment_density_score,
            'Readability Score'     : r.readability_score,
            'Category'              : r.category,
        } for r in self.results])

        df = df.sort_values('Readability Score')

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

            # Sheet 1 – Results
            df.to_excel(writer, sheet_name='Results', index=False)
            ws = writer.sheets['Results']
            hdr_fill = PatternFill('solid', fgColor='366092')
            hdr_font = Font(color='FFFFFF', bold=True)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                w = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(w + 2, 50)

            # Sheet 2 – Summary
            self._summary_df().to_excel(writer, sheet_name='Summary', index=True)

            # Sheet 3 – Module Comparison
            comp = df.groupby('Module').agg(
                Fungsi=('Readability Score', 'count'),
                Rata_Skor=('Readability Score', 'mean'),
                Median_Skor=('Readability Score', 'median'),
                Min_Skor=('Readability Score', 'min'),
                Max_Skor=('Readability Score', 'max'),
                Std_Skor=('Readability Score', 'std'),
                Rata_LOC=('Code Lines (LOC)', 'mean'),
                Rata_CD=('Comment Density (%)', 'mean'),
                Pct_High=('Category', lambda x: round((x == 'High').mean() * 100, 2)),
                Pct_Medium=('Category', lambda x: round((x == 'Medium').mean() * 100, 2)),
                Pct_Low=('Category', lambda x: round((x == 'Low').mean() * 100, 2)),
            ).round(2)
            comp.to_excel(writer, sheet_name='Module Comparison')

            # Sheet 4 – Low Readability
            df[df['Category'] == 'Low'].to_excel(
                writer, sheet_name='Low Readability', index=False)

        # ── Tulis formula Category & warna dengan openpyxl (setelah file disimpan) ──
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb['Results']
        color_map = {'High': 'C6EFCE', 'Medium': 'FFEB9C', 'Low': 'FFC7CE'}
        for i in range(2, ws.max_row + 1):
            rs = ws.cell(row=i, column=14).value   # Readability Score kolom N
            cat_cell = ws.cell(row=i, column=15)   # Category kolom O
            cat_cell.value = f'=IF(N{i}>=80,"High",IF(N{i}>=50,"Medium","Low"))'
            if rs is not None:
                if rs >= 80:
                    cat_cell.fill = PatternFill('solid', fgColor=color_map['High'])
                elif rs >= 50:
                    cat_cell.fill = PatternFill('solid', fgColor=color_map['Medium'])
                else:
                    cat_cell.fill = PatternFill('solid', fgColor=color_map['Low'])
        wb.save(output_file)
        wb.close()

        print(f"✓ Excel berhasil dibuat: {output_file}")

    def _summary_df(self) -> pd.DataFrame:
        df = pd.DataFrame([vars(r) for r in self.results])
        data = {
            'Total Fungsi Dianalisis'    : len(df),
            'Rata-rata Readability Score': round(df['readability_score'].mean(), 2),
            'Median Readability Score'   : round(df['readability_score'].median(), 2),
            'Std Dev Readability Score'  : round(df['readability_score'].std(), 2),
            'Jumlah High'                : int((df['category'] == 'High').sum()),
            'Jumlah Medium'              : int((df['category'] == 'Medium').sum()),
            'Jumlah Low'                 : int((df['category'] == 'Low').sum()),
            '% High'                     : round((df['category'] == 'High').mean() * 100, 2),
            '% Medium'                   : round((df['category'] == 'Medium').mean() * 100, 2),
            '% Low'                      : round((df['category'] == 'Low').mean() * 100, 2),
            'Rata-rata LOC'              : round(df['code_lines'].mean(), 2),
            'Median LOC'                 : round(df['code_lines'].median(), 2),
            'Fungsi > 50 LOC'            : int((df['code_lines'] > 50).sum()),
            '% Fungsi > 50 LOC'          : round((df['code_lines'] > 50).mean() * 100, 2),
            'Rata-rata Comment Density %': round(df['comment_density'].mean(), 2),
            'Median Comment Density %'   : round(df['comment_density'].median(), 2),
            'Fungsi CD < 10%'            : int((df['comment_density'] < 10).sum()),
            '% Fungsi CD < 10%'          : round((df['comment_density'] < 10).mean() * 100, 2),
        }
        return pd.DataFrame.from_dict(data, orient='index', columns=['Value'])

    # ── Visualisasi ──────────────────────────────────────────────────────────
    def generate_visualizations(self, viz_dir: str = 'visualizations') -> None:
        print(f"Membuat visualisasi di folder '{viz_dir}/'...")
        out = Path(viz_dir)
        out.mkdir(exist_ok=True)

        df = pd.DataFrame([vars(r) for r in self.results])

        # 1. Distribusi Readability Score
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.hist(df['readability_score'], bins=20, edgecolor='black', alpha=0.7, color='steelblue')
        ax.axvline(df['readability_score'].mean(),   color='red',   linestyle='--',
                   label=f'Mean: {df["readability_score"].mean():.2f}')
        ax.axvline(df['readability_score'].median(), color='green', linestyle='--',
                   label=f'Median: {df["readability_score"].median():.2f}')
        ax.set(xlabel='Readability Score', ylabel='Frekuensi',
               title='Distribusi Readability Score')
        ax.legend(); ax.grid(True, alpha=0.3)
        fig.savefig(out / '1_readability_distribution.png', dpi=300, bbox_inches='tight')
        plt.close(fig)

        # 2. Rata-rata Skor per Modul
        fig, ax = plt.subplots(figsize=(14, 7))
        mod_scores = df.groupby('module')['readability_score'].mean().sort_values()
        colors = ['#d32f2f' if x < 50 else '#ffa726' if x < 80 else '#66bb6a'
                  for x in mod_scores]
        mod_scores.plot(kind='barh', color=colors, ax=ax)
        ax.axvline(50, color='orange', linestyle='--', alpha=0.6, label='Batas Medium (50)')
        ax.axvline(80, color='green',  linestyle='--', alpha=0.6, label='Batas High (80)')
        ax.set(xlabel='Rata-rata Readability Score', ylabel='Modul',
               title='Rata-rata Readability Score per Modul')
        ax.legend(); ax.grid(True, alpha=0.3, axis='x')
        fig.tight_layout()
        fig.savefig(out / '2_readability_by_module.png', dpi=300, bbox_inches='tight')
        plt.close(fig)

        # 3. Pie Chart Kategori
        fig, ax = plt.subplots(figsize=(8, 8))
        counts = df['category'].value_counts()
        ax.pie(counts, labels=counts.index, autopct='%1.1f%%', startangle=90,
               colors=['#66bb6a', '#ffa726', '#d32f2f'])
        ax.set_title('Distribusi Kategori Readability')
        fig.savefig(out / '3_category_distribution.png', dpi=300, bbox_inches='tight')
        plt.close(fig)

        # 4. Scatter LOC vs Comment Density
        fig, ax = plt.subplots(figsize=(10, 7))
        sc = ax.scatter(df['code_lines'], df['comment_density'],
                        c=df['readability_score'], cmap='RdYlGn', alpha=0.5, s=40)
        plt.colorbar(sc, ax=ax, label='Readability Score')
        ax.axvline(50, color='red',    linestyle='--', alpha=0.4, label='LOC > 50')
        ax.axhline(10, color='orange', linestyle='--', alpha=0.4, label='CD < 10%')
        ax.axhline(25, color='green',  linestyle='--', alpha=0.4, label='CD ideal 25%')
        ax.set(xlabel='Lines of Code (LOC)', ylabel='Comment Density (%)',
               title='Panjang Fungsi vs Comment Density')
        ax.legend(); ax.grid(True, alpha=0.3)
        fig.savefig(out / '4_loc_vs_comments.png', dpi=300, bbox_inches='tight')
        plt.close(fig)

        # 5. Box Plot per Modul
        fig, ax = plt.subplots(figsize=(14, 6))
        modules_sorted = df.groupby('module')['readability_score'].median().sort_values().index
        data_plot = [df[df['module'] == m]['readability_score'].values for m in modules_sorted]
        ax.boxplot(data_plot, labels=modules_sorted, patch_artist=True)
        ax.set(xlabel='Modul', ylabel='Readability Score',
               title='Variasi Readability Score per Modul')
        plt.xticks(rotation=45, ha='right')
        ax.grid(True, alpha=0.3)
        fig.tight_layout()
        fig.savefig(out / '5_module_boxplot.png', dpi=300, bbox_inches='tight')
        plt.close(fig)

        # 6. Distribusi Panjang Fungsi (LOC)
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.hist(df['code_lines'], bins=30, edgecolor='black', alpha=0.7, color='coral')
        ax.axvline(20, color='green',  linestyle='--', label='Ideal (20 LOC)')
        ax.axvline(50, color='orange', linestyle='--', label='Batas wajar (50 LOC)')
        ax.axvline(df['code_lines'].mean(), color='red', linestyle='--',
                   label=f'Mean: {df["code_lines"].mean():.1f}')
        ax.set(xlabel='Lines of Code (LOC)', ylabel='Frekuensi',
               title='Distribusi Panjang Fungsi')
        ax.legend(); ax.grid(True, alpha=0.3)
        fig.savefig(out / '6_function_length_distribution.png', dpi=300, bbox_inches='tight')
        plt.close(fig)

        # 7. Distribusi Comment Density
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.hist(df['comment_density'], bins=30, edgecolor='black', alpha=0.7,
                color='mediumseagreen')
        ax.axvline(10, color='orange', linestyle='--', label='Minimum (10%)')
        ax.axvline(25, color='green',  linestyle='--', label='Ideal (25%)')
        ax.axvline(df['comment_density'].mean(), color='red', linestyle='--',
                   label=f'Mean: {df["comment_density"].mean():.1f}%')
        ax.set(xlabel='Comment Density (%)', ylabel='Frekuensi',
               title='Distribusi Comment Density')
        ax.legend(); ax.grid(True, alpha=0.3)
        fig.savefig(out / '7_comment_density_distribution.png', dpi=300, bbox_inches='tight')
        plt.close(fig)

        # 8. Stacked Bar: proporsi kategori per modul
        fig, ax = plt.subplots(figsize=(14, 7))
        cat_pct = (df.groupby('module')['category']
                   .value_counts(normalize=True)
                   .unstack(fill_value=0) * 100)
        for col in ['High', 'Medium', 'Low']:
            if col not in cat_pct.columns:
                cat_pct[col] = 0
        cat_pct = cat_pct[['High', 'Medium', 'Low']]
        cat_pct.sort_values('High').plot(
            kind='barh', stacked=True, ax=ax,
            color=['#66bb6a', '#ffa726', '#d32f2f'])
        ax.set(xlabel='Persentase (%)', ylabel='Modul',
               title='Proporsi Kategori Readability per Modul')
        ax.legend(title='Kategori', loc='lower right')
        ax.grid(True, alpha=0.3, axis='x')
        fig.tight_layout()
        fig.savefig(out / '8_category_stacked_by_module.png', dpi=300, bbox_inches='tight')
        plt.close(fig)

        print(f"✓ 8 visualisasi tersimpan di '{viz_dir}/'")


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='Moodle Readability Analyzer v3.0',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Contoh penggunaan:

  # Scan SEMUA folder, TANPA sampling (semua file diproses — bisa lama!)
  python moodle_readability_analyzer.py --path "C:/Users/user/Downloads/moodle-latest-501 (1)/moodle/public"

  # Scan modul tertentu, 100 file per modul (sampling aktif)
  python moodle_readability_analyzer.py --path "..." --modules lib mod search --sample 100

  # Scan modul tertentu, tanpa sampling
  python moodle_readability_analyzer.py --path "..." --modules lib mod
        """
    )
    parser.add_argument('--path',    '-p', required=True,
                        help='Path ke folder moodle/public')
    parser.add_argument('--modules', '-m', nargs='+', default=None,
                        help='Modul yang dianalisis (default: semua subfolder otomatis)')
    parser.add_argument('--sample',  '-s', type=int, default=None,
                        help='Jumlah FILE yang disampling acak per modul (contoh: --sample 100)')
    parser.add_argument('--seed',         type=int, default=42,
                        help='Random seed untuk reproduksibilitas (default: 42)')
    parser.add_argument('--output',  '-o', default='moodle_readability_results.xlsx',
                        help='Nama file Excel output')
    parser.add_argument('--viz-dir', '-v', default='visualizations',
                        help='Folder untuk menyimpan grafik')
    parser.add_argument('--no-viz',  action='store_true',
                        help='Lewati pembuatan grafik')

    args = parser.parse_args()

    try:
        analyzer = MoodleReadabilityAnalyzer(
            moodle_path = args.path,
            modules     = args.modules,
            sample_size = args.sample,
            seed        = args.seed,
        )

        analyzer.analyze()

        if not analyzer.results:
            print("⚠ Tidak ada fungsi yang ditemukan. Periksa path Moodle.")
            return

        analyzer.export_excel(args.output)

        if not args.no_viz:
            analyzer.generate_visualizations(args.viz_dir)

        print("\n" + "=" * 60)
        print("RINGKASAN STATISTIK")
        print("=" * 60)
        print(analyzer._summary_df().to_string())
        print("=" * 60)
        print(f"\n✓ Selesai!")
        print(f"  📊 Excel  : {args.output}")
        if not args.no_viz:
            print(f"  📈 Grafik : {args.viz_dir}/")

    except FileNotFoundError as e:
        print(f"\n❌ Error: {e}")
        sys.exit(1)
    except Exception as e:
        import traceback
        print(f"\n❌ Error tidak terduga: {e}")
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()